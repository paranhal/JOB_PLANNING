#!/usr/bin/env python3
"""
엑셀 업무일지 파일의 모든 시트를 분석하여 결과를 마크다운으로 저장합니다.
사용법:
  python3 analyze_excel.py
    → 프로젝트 루트에서 *.xlsx 파일 찾아 분석
  python3 analyze_excel.py "/경로/파일.xlsx"
    → 지정한 엑셀 파일 분석
결과: 엑셀_시트_분석결과.md
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

def find_xlsx_in_project():
    root = Path(__file__).resolve().parent
    for f in root.glob("*.xlsx"):
        return str(f)
    return None

def main():
    if len(sys.argv) >= 2:
        path = Path(sys.argv[1])
    else:
        path = find_xlsx_in_project()
        if not path:
            print("사용법: python3 analyze_excel.py <엑셀파일경로>")
            print("또는 프로젝트 루트에 .xlsx 파일을 넣고 python3 analyze_excel.py")
            sys.exit(1)
        path = Path(path)

    if not path.exists():
        print(f"파일 없음: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    out_lines = [
        "# 엑셀 업무일지 시트 분석 결과",
        "",
        f"**파일:** `{path.name}`",
        "",
        "## 시트 목록",
        "",
        "| 순번 | 시트명 |",
        "|------|--------|",
    ]
    for i, name in enumerate(wb.sheetnames, 1):
        out_lines.append(f"| {i} | {name} |")

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        out_lines.extend([
            "",
            "---",
            "",
            f"## 시트: {name}",
            "",
            f"- **행 수:** {len(rows)}",
            f"- **열 수:** {len(rows[0]) if rows else 0}",
            "",
            "### 컬럼(첫 행) 및 샘플 데이터",
            "",
        ])
        if not rows:
            out_lines.append("(데이터 없음)")
            continue
        # 첫 행을 헤더로
        headers = [str(c) if c is not None else "" for c in rows[0]]
        out_lines.append("| " + " | ".join(headers) + " |")
        out_lines.append("| " + " | ".join("---" for _ in headers) + " |")
        for row in rows[1:11]:  # 최대 10행 샘플
            cells = [str(c)[:50] if c is not None else "" for c in row]
            while len(cells) < len(headers):
                cells.append("")
            out_lines.append("| " + " | ".join(cells[:len(headers)]) + " |")
        if len(rows) > 11:
            out_lines.append(f"| ... 외 {len(rows)-11}행 |")
        out_lines.append("")
        # 컬럼별 정리
        out_lines.append("### 컬럼 목록 (입력/관리 항목 후보)")
        out_lines.append("")
        for j, h in enumerate(headers):
            if h and str(h).strip():
                out_lines.append(f"- `{j+1}` {h}")
        out_lines.append("")

    wb.close()

    out_path = Path(__file__).resolve().parent / "엑셀_시트_분석결과.md"
    out_path.write_text("\n".join(out_lines), encoding="utf-8")
    print(f"분석 완료 → {out_path}")

if __name__ == "__main__":
    main()
