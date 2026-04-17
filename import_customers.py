#!/usr/bin/env python3
"""
기존 엑셀 파일에서 고객명과 담당자 정보를 추출하여 SQLite DB에 삽입합니다.

대상 시트:
  - NICOM_유지보수현황        (No | 고객명 | 구분 | 담당자 | 직장 | 핸드폰 | 이메일 | ...)
  - VISIONIT_MaintenanceList  (No | 고객명 | 구분 | 담당자 | 직장 | 핸드폰 | 이메일 | ...)
  - 충남교육청_KLAS            (No | 고객명 | 담당자 | 직장 | 핸드폰 | 이메일 | ...)
  - 세종시도서관_KLAS          (No | 고객명 | 담당자 | 직장 | 핸드폰 | 이메일 | ...)

사용법:
  python import_customers.py --dry-run     # 미리보기 (DB 변경 없음)
  python import_customers.py               # 실제 삽입

  python import_customers.py "C:/경로/파일.xlsx" --dry-run   # 파일 지정
"""
import sys
import sqlite3
import uuid
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "server" / "data" / "app.db"

# ── 시트별 컬럼 설정 ──────────────────────────────────────────────────────────
# header_row: 실제 컬럼 헤더가 있는 행 번호 (0-indexed)
# cols: 각 필드가 몇 번째 열(0-indexed)인지
SHEET_CONFIGS = {
    "NICOM_유지보수현황": {
        "header_row": 1,
        "cols": {"고객명": 1, "담당자": 3, "직장": 4, "핸드폰": 5, "이메일": 6},
    },
    "VISIONIT_MaintenanceList": {
        "header_row": 1,
        "cols": {"고객명": 1, "담당자": 3, "직장": 4, "핸드폰": 5, "이메일": 6},
    },
    "충남교육청_KLAS": {
        "header_row": 1,
        "cols": {"고객명": 1, "담당자": 2, "직장": 3, "핸드폰": 4, "이메일": 5},
    },
    "세종시도서관_KLAS": {
        "header_row": 1,
        "cols": {"고객명": 1, "담당자": 2, "직장": 3, "핸드폰": 4, "이메일": 5},
    },
}

# 헤더 행에서 나타나는 값(실제 데이터가 아닌 것) – 스킵
SKIP_VALUES = {"no.", "고객명", "no", "번호", "none", ""}


def gen_id(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8].upper()}"


def cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # 날짜가 섞여 들어오는 경우 방지
    if "00:00:00" in s:
        return ""
    return s


def find_excel(root: Path):
    for f in root.glob("*.xlsx"):
        return f
    return None


def extract_data(excel_path: Path):
    """엑셀 파일에서 고객·담당자 데이터 추출."""
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    customers: dict[str, dict] = {}   # org_name → customer dict
    contacts: dict[tuple, dict] = {}  # (org_name, full_name) → contact dict

    for sheet_name, config in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [경고] 시트 없음: {sheet_name} (스킵)")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = config["header_row"]
        cols = config["cols"]

        prev_org_name = ""
        count = 0

        for row in rows[header_row + 1:]:
            # 완전히 빈 행 스킵
            if not any(c for c in row):
                continue

            def get(col_key: str) -> str:
                idx = cols.get(col_key)
                if idx is None or idx >= len(row):
                    return ""
                return cell_str(row[idx])

            org_name = get("고객명")
            # 병합 셀 대응: 고객명이 비어 있으면 이전 값 사용
            if not org_name:
                org_name = prev_org_name
            else:
                prev_org_name = org_name

            # 헤더 값 또는 빈 값이면 스킵
            if org_name.lower() in SKIP_VALUES:
                continue

            contact_name = get("담당자")
            phone        = get("직장")
            mobile       = get("핸드폰")
            email        = get("이메일")

            # 고객 등록 (중복 방지)
            if org_name not in customers:
                customers[org_name] = {
                    "customer_id":   gen_id("CUST"),
                    "org_name":      org_name,
                    "official_name": org_name,
                    "industry":      "도서관",
                    "is_active":     True,
                }

            # 담당자 등록 (이름이 있을 때만, 중복 방지)
            if contact_name and contact_name.lower() not in SKIP_VALUES and contact_name != "담당자":
                key = (org_name, contact_name)
                if key not in contacts:
                    contacts[key] = {
                        "contact_id":  gen_id("CONT"),
                        "org_name":    org_name,
                        "full_name":   contact_name,
                        "phone":       phone,
                        "mobile":      mobile,
                        "email":       email,
                        "status":      "active",
                        "is_primary":  True,
                    }
                else:
                    # 이미 등록된 담당자라면 비어 있는 연락처만 보완
                    existing = contacts[key]
                    if not existing["phone"] and phone:
                        existing["phone"] = phone
                    if not existing["mobile"] and mobile:
                        existing["mobile"] = mobile
                    if not existing["email"] and email:
                        existing["email"] = email

            count += 1

        print(f"  [{sheet_name}] {count}행 처리 → "
              f"누계 고객 {len(customers)}개 / 담당자 {len(contacts)}개")

    wb.close()
    return list(customers.values()), list(contacts.values())


def print_preview(customers: list, contacts: list):
    print(f"\n{'='*60}")
    print(f"[미리보기] 고객 {len(customers)}개, 담당자 {len(contacts)}개 삽입 예정")
    print(f"{'='*60}")

    print(f"\n■ 고객 목록 (총 {len(customers)}개)")
    for i, c in enumerate(customers, 1):
        print(f"  {i:3d}. {c['org_name']}")

    print(f"\n■ 담당자 목록 (총 {len(contacts)}개)")
    for i, c in enumerate(contacts, 1):
        parts = [p for p in [c["phone"], c["mobile"], c["email"]] if p]
        print(f"  {i:3d}. [{c['org_name']}] {c['full_name']}"
              + (f"  ←  {' / '.join(parts)}" if parts else ""))

    print(f"\n실제 삽입하려면 --dry-run 옵션 없이 다시 실행하세요.")


def insert_to_db(db_path: Path, customers: list, contacts: list):
    now = datetime.now().isoformat()
    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    # ── 고객 삽입 ──────────────────────────────────────────────────────────────
    cust_inserted = 0
    cust_skipped  = 0
    org_to_cid: dict[str, str] = {}

    for c in customers:
        cur.execute(
            "SELECT customer_id FROM customers WHERE org_name = ?",
            (c["org_name"],)
        )
        row = cur.fetchone()
        if row:
            org_to_cid[c["org_name"]] = row[0]
            cust_skipped += 1
        else:
            cur.execute("""
                INSERT INTO customers
                    (customer_id, org_name, official_name, industry, is_active,
                     created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                c["customer_id"], c["org_name"], c["official_name"],
                c["industry"], 1, now, now
            ))
            org_to_cid[c["org_name"]] = c["customer_id"]
            cust_inserted += 1

    # ── 담당자 삽입 ────────────────────────────────────────────────────────────
    cont_inserted = 0
    cont_skipped  = 0

    for c in contacts:
        customer_id = org_to_cid.get(c["org_name"])
        if not customer_id:
            continue

        cur.execute(
            "SELECT contact_id FROM contacts WHERE customer_id = ? AND full_name = ?",
            (customer_id, c["full_name"])
        )
        if cur.fetchone():
            cont_skipped += 1
        else:
            cur.execute("""
                INSERT INTO contacts
                    (contact_id, customer_id, full_name, phone, mobile, email,
                     status, is_primary, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                c["contact_id"], customer_id, c["full_name"],
                c["phone"], c["mobile"], c["email"],
                c["status"], 1 if c["is_primary"] else 0, now
            ))
            cont_inserted += 1

    conn.commit()
    conn.close()

    print(f"\n{'='*60}")
    print(f"[완료] 고객  삽입 {cust_inserted:3d}개  /  중복 스킵 {cust_skipped:3d}개")
    print(f"[완료] 담당자 삽입 {cont_inserted:3d}개  /  중복 스킵 {cont_skipped:3d}개")
    print(f"{'='*60}")


def main():
    dry_run = "--dry-run" in sys.argv or "-n" in sys.argv

    # 엑셀 파일 경로 결정
    excel_path: Path | None = None
    for arg in sys.argv[1:]:
        if not arg.startswith("-"):
            excel_path = Path(arg)
            break
    if excel_path is None:
        excel_path = find_excel(ROOT)

    if excel_path is None or not excel_path.exists():
        print("엑셀 파일을 찾을 수 없습니다.")
        print("사용법: python import_customers.py [엑셀파일] [--dry-run]")
        sys.exit(1)

    if not DB_PATH.exists():
        print(f"DB 파일이 없습니다: {DB_PATH}")
        sys.exit(1)

    print(f"엑셀  : {excel_path.name}")
    print(f"DB    : {DB_PATH}")
    print(f"모드  : {'미리보기 (dry-run) - DB 변경 없음' if dry_run else '실제 삽입'}")
    print()

    print("데이터 추출 중...")
    customers, contacts = extract_data(excel_path)

    if dry_run:
        print_preview(customers, contacts)
    else:
        insert_to_db(DB_PATH, customers, contacts)


if __name__ == "__main__":
    main()
